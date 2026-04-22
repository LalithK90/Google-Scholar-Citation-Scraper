import json
import logging
from pathlib import Path
from typing import Any, Dict, List

try:
    from .utils import sanitize_filename, unique_sheet_name
except ImportError:
    from utils import sanitize_filename, unique_sheet_name


def _build_rows(scraper: Any):
    """Return (rows_json_mode, main_rows, cited_rows, per_pub_sheets)."""
    rows_json_mode: List[Dict] = []
    main_rows: List[Dict] = []
    cited_rows: List[Dict] = []
    per_pub_sheets: Dict[str, List[Dict]] = {}
    max_authors = 0
    for pub in scraper.data.get("publications", []):
        details = pub.get("details") or {}
        authors = details.get("authors") or []
        max_authors = max(max_authors, len(authors))

    for pub in scraper.data.get("publications", []):
        details = pub.get("details") or {}
        authors = details.get("authors") or []
        rows_json_mode.append(
            {
                "no": pub.get("no"),
                "title": pub.get("title"),
                "paper_details_link": pub.get("paper_details_link"),
                "citation_count": pub.get("citation_count"),
                "citation_link": pub.get("citation_link"),
                "year": pub.get("year"),
                "details": json.dumps(pub.get("details"), ensure_ascii=False),
                "citations": json.dumps(pub.get("citations"), ensure_ascii=False),
            }
        )
        row = {
            "no": pub.get("no"),
            "title": pub.get("title"),
            "paper_details_link": pub.get("paper_details_link"),
            "citation_count": pub.get("citation_count"),
            "citation_link": pub.get("citation_link"),
            "year": pub.get("year"),
            "publication_date": details.get("publication_date"),
            "journal": details.get("journal"),
            "volume": details.get("volume"),
            "issue": details.get("issue"),
            "pages": details.get("pages"),
        }
        for i in range(max_authors):
            row[f"author_{i+1}"] = authors[i] if i < len(authors) else ""
        main_rows.append(row)

        for ci, cited in enumerate(pub.get("citations", []) or [], start=1):
            cited_row = {
                "parent_no": pub.get("no"),
                "parent_title": pub.get("title"),
                "cited_no": ci,
                "cited_paper_name": cited.get("cited_paper_name"),
                "cited_paper_link": cited.get("cited_paper_link"),
                "journal_name": cited.get("journal_name"),
                "doi": cited.get("doi"),
                "is_duplicate": cited.get("is_duplicate", False),
            }
            for j in range(5):
                cited_row[f"cited_author_{j+1}"] = (
                    cited.get("cited_authors")[j] if j < len(cited.get("cited_authors", [])) else ""
                )
            cited_rows.append(cited_row)

            pub_no = pub.get("no") or ""
            pub_cites = pub.get("citation_count") or 0
            pub_year = pub.get("year") or ""
            sheet_title = sanitize_filename(f"{pub_no}_{pub_cites}_{pub_year}")
            if pub_cites > 0:
                per_rows = per_pub_sheets.setdefault(sheet_title, [])
                per_rows.append(cited_row)

    return rows_json_mode, main_rows, cited_rows, per_pub_sheets


def export_excel(scraper: Any, stream: bool = False) -> None:
    """Export the scraper.data to an Excel workbook.

    If stream is True, use openpyxl write-only mode to keep memory usage low.
    Otherwise the function uses pandas/openpyxl.
    """
    if not getattr(scraper, "author_sanitized", None):
        logging.error("Author metadata not extracted; cannot export Excel")
        return

    excel_filename = getattr(scraper, "excel_path", None)
    if not excel_filename:
        output_dir = Path(getattr(scraper, "download_dir", "."))
        excel_filename = str(output_dir / f"{scraper.author_sanitized}.xlsx")
    logging.info(f"Exporting Excel to {excel_filename} in mode {getattr(scraper, 'excel_mode', 'flat')} (stream={stream})")

    rows_json_mode, main_rows, cited_rows, per_pub_sheets = _build_rows(scraper)

    # Build validation rows
    validation_list = scraper.data.get("validation", {}).get("citation_mismatches", []) or []
    val_rows = []
    for v in validation_list:
        reported = int(v.get("reported_citation_count") or 0)
        saved = int(v.get("saved_citations_len") or 0)
        missing = reported - saved
        if missing > 0:
            action = f"re-fetch citations (missing {missing})"
        elif missing == 0:
            action = "ok"
        else:
            action = "investigate (more saved than reported)"
        val_rows.append(
            {
                "no": v.get("no"),
                "title": v.get("title"),
                "reported_citation_count": reported,
                "saved_citations_len": saved,
                "missing_count": missing,
                "suggested_action": action,
            }
        )

    if stream:
        try:
            from openpyxl import Workbook
        except Exception:
            logging.error("openpyxl is required for streaming Excel export. Install via requirements.txt")
            return

        wb = Workbook(write_only=True)
        # Cover sheet
        ws_cover = wb.create_sheet("cover")
        author = scraper.data.get("profile", {})
        summary = scraper.data.get("analysis", {}) or {}
        kv = [
            ("author_name", author.get("author_name")),
            ("affiliation", author.get("affiliation")),
            ("profile_url", author.get("profile_url")),
            ("last_updated", scraper.data.get("last_updated")),
            ("total_publications", len(scraper.data.get("publications", []))),
            ("h_index_deduped", summary.get("h_index_deduped")),
        ]
        for k, v in kv:
            ws_cover.append([k, v])

        # Main publications sheet
        ws_main = wb.create_sheet(sanitized := (scraper.author_sanitized[:31]))
        main_headers = [
            "no",
            "title",
            "paper_details_link",
            "citation_count",
            "citation_link",
            "year",
            "publication_date",
            "journal",
            "volume",
            "issue",
            "pages",
        ]
        ws_main.append(main_headers)
        for pub in scraper.data.get("publications", []):
            details = pub.get("details") or {}
            row = [
                pub.get("no"),
                pub.get("title"),
                pub.get("paper_details_link"),
                pub.get("citation_count"),
                pub.get("citation_link"),
                pub.get("year"),
                details.get("publication_date"),
                details.get("journal"),
                details.get("volume"),
                details.get("issue"),
                details.get("pages"),
            ]
            ws_main.append(row)

        # cited_papers sheet
        ws_cited = wb.create_sheet("cited_papers")
        ws_cited.append([
            "parent_no",
            "parent_title",
            "cited_no",
            "cited_paper_name",
            "cited_paper_link",
            "journal_name",
            "doi",
            "is_duplicate",
        ])
        for pub in scraper.data.get("publications", []):
            for ci, cited in enumerate(pub.get("citations", []) or [], start=1):
                ws_cited.append([
                    pub.get("no"),
                    pub.get("title"),
                    ci,
                    cited.get("cited_paper_name"),
                    cited.get("cited_paper_link"),
                    cited.get("journal_name"),
                    cited.get("doi"),
                    cited.get("is_duplicate", False),
                ])

        # Per-cited sheets (if requested)
        if getattr(scraper, "per_cited_sheet", False):
            per_cited = {}
            for pub in scraper.data.get("publications", []):
                for ci, cited in enumerate(pub.get("citations", []) or [], start=1):
                    key = cited.get("doi") or cited.get("cited_paper_link") or cited.get("cited_paper_name")
                    if not key:
                        continue
                    sn = sanitize_filename(str(key))[:28]
                    per_cited.setdefault(sn, []).append(cited)
            for sn, items in per_cited.items():
                ws = wb.create_sheet(f"c_{sn}")
                ws.append(["parent_no", "cited_paper_name", "cited_paper_link", "journal_name", "doi", "is_duplicate"])
                for cited in items:
                    ws.append([
                        cited.get("parent_no", ""),
                        cited.get("cited_paper_name"),
                        cited.get("cited_paper_link"),
                        cited.get("journal_name"),
                        cited.get("doi"),
                        cited.get("is_duplicate", False),
                    ])

        wb.save(excel_filename)
        logging.info(f"Streaming Excel export completed: {excel_filename}")
        return

    # Non-stream path (pandas)
    try:
        import pandas as pd
    except Exception:
        logging.error("pandas is required for non-stream export. Install via requirements.txt")
        return

    # sanitized sheet name for non-stream path (pandas)
    sanitized = sanitize_filename(getattr(scraper, "author_sanitized", ""))[:31]

    df_json_main = pd.DataFrame(rows_json_mode)
    df_flat_main = pd.DataFrame(main_rows)
    df_cited = pd.DataFrame(cited_rows)

    # Build key/value summary and write via pandas ExcelWriter
    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        sheet_name_val = unique_sheet_name(writer, "validation_report")
        sheet_name_sum = unique_sheet_name(writer, "validation_summary")
        sheet_name_cover = unique_sheet_name(writer, "cover")

        author = scraper.data.get("profile", {})
        summary = scraper.data.get("analysis", {}) or {}
        total_pubs = len(scraper.data.get("publications", []))
        total_mismatches = len(val_rows)
        total_missing = sum([r.get("missing_count", 0) for r in val_rows])
        kv = [
            {"key": "author_name", "value": author.get("author_name")},
            {"key": "affiliation", "value": author.get("affiliation")},
            {"key": "profile_url", "value": author.get("profile_url")},
            {"key": "last_updated", "value": scraper.data.get("last_updated")},
            {"key": "total_publications", "value": total_pubs},
            {"key": "h_index_deduped", "value": summary.get("h_index_deduped")},
            {"key": "total_citations_reported", "value": summary.get("total_citations_reported")},
            {"key": "unique_citations", "value": summary.get("unique_citations")},
            {"key": "duplicate_count", "value": summary.get("duplicate_count")},
            {"key": "total_mismatches", "value": total_mismatches},
            {"key": "total_missing_citations", "value": total_missing},
        ]
        df_kv = pd.DataFrame(kv)

        mismatch_rows = []
        for idx, v in enumerate(val_rows):
            rownum = idx + 2
            link_formula = f'=HYPERLINK("#\'{sheet_name_val}\'!A{rownum}", "View")'
            mismatch_rows.append(
                {
                    "no": v.get("no"),
                    "title": v.get("title"),
                    "reported_citation_count": v.get("reported_citation_count"),
                    "saved_citations_len": v.get("saved_citations_len"),
                    "missing_count": v.get("missing_count"),
                    "link": link_formula,
                }
            )
        df_mismatches = pd.DataFrame(mismatch_rows)

        df_kv.to_excel(writer, index=False, sheet_name=sheet_name_cover, startrow=0)
        startrow = len(df_kv) + 2
        if not df_mismatches.empty:
            df_mismatches.to_excel(writer, index=False, sheet_name=sheet_name_cover, startrow=startrow)

        if getattr(scraper, "excel_mode", "flat") == "json_column":
            df_json_main.to_excel(writer, index=False, sheet_name=sanitized)
        else:
            df_flat_main.to_excel(writer, index=False, sheet_name=sanitized)

        for sheet_name, rows_list in per_pub_sheets.items():
            try:
                df_per = pd.DataFrame(rows_list)
                sheet_nm = unique_sheet_name(writer, sheet_name)
                df_per.to_excel(writer, index=False, sheet_name=sheet_nm)
            except Exception:
                logging.exception(f"Failed to write per-publication sheet {sheet_name}")

        if len(cited_rows) > 0:
            df_cited.to_excel(writer, index=False, sheet_name="cited_papers")

        try:
            sheet_name_val = unique_sheet_name(writer, "validation_report")
            df_validation = pd.DataFrame(val_rows)
            df_validation.to_excel(writer, index=False, sheet_name=sheet_name_val)
        except Exception:
            logging.exception("Failed to write validation_report sheet")

        summary_rows = [
            {
                "total_publications": total_pubs,
                "total_mismatches": total_mismatches,
                "total_missing_citations": total_missing,
            }
        ]
        df_summary = pd.DataFrame(summary_rows)
        try:
            sheet_name_sum = unique_sheet_name(writer, "validation_summary")
            df_summary.to_excel(writer, index=False, sheet_name=sheet_name_sum)
        except Exception:
            logging.exception("Failed to write validation_summary sheet")

        if getattr(scraper, "per_cited_sheet", False) and len(cited_rows) > 0:
            per_cited = {}
            for r in cited_rows:
                key = r.get("doi") or r.get("cited_paper_link") or r.get("cited_paper_name")
                if not key:
                    continue
                sn = sanitize_filename(str(key))[:28]
                per_cited.setdefault(sn, []).append(r)
            for sn, rows_list in per_cited.items():
                try:
                    df_sn = pd.DataFrame(rows_list)
                    sheet_nm = unique_sheet_name(writer, f"c_{sn}")
                    df_sn.to_excel(writer, index=False, sheet_name=sheet_nm)
                except Exception:
                    logging.exception(f"Failed to write per-cited sheet {sn}")

    logging.info(f"Excel export completed: {excel_filename}")
